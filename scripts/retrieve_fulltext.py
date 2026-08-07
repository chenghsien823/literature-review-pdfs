#!/usr/bin/env python3
"""Retrieve legally accessible full-text PDFs for bibliographic records.

This script intentionally supports open sources only.  --browser-session does
not read browser data or log in; it marks publisher pages for a human handoff.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

EUTILS = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
EUROPE_PMC = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
UNPAYWALL = "https://api.unpaywall.org/v2/"
USER_AGENT = "retrieve-journal-pdfs/1.0 (legal-open-access-retrieval)"
DETAIL_HEADERS = [
    "record_id", "citation", "pmid", "doi", "title", "retrieval_status",
    "source", "access_type", "candidate_url", "local_path_or_url",
    "content_type", "sha256", "page_count", "notes", "next_action",
    "retrieved_utc", "first_author_country", "filename_style",
]
SRMA_HEADERS = [
    "record_id", "citation", "pmid", "doi", "retrieval_status", "source",
    "local_path_or_url", "notes",
]


def text(value: Any) -> str:
    return str(value or "").strip()


def normalize_doi(value: Any) -> str:
    value = text(value).lower()
    value = re.sub(r"^https?://(dx\.)?doi\.org/", "", value)
    return value.rstrip(" .;,")


def normalize_pmid(value: Any) -> str:
    match = re.search(r"\d{4,10}", text(value))
    return match.group(0) if match else ""


def norm_title(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", text(value).lower())


def clean_component(value: Any, default: str, max_len: int = 36) -> str:
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", text(value))
    value = re.sub(r"\s+", "_", value).strip("._ ")
    return (value or default)[:max_len]


def clean_label(value: Any, default: str, max_len: int = 36) -> str:
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', " ", text(value))
    value = re.sub(r"\s+", " ", value).strip("._ ")
    return (value or default)[:max_len]


def record_key(record: dict[str, str]) -> str:
    if record["doi"]:
        return "doi:" + record["doi"]
    if record["pmid"]:
        return "pmid:" + record["pmid"]
    return "title:" + norm_title(record["title"])


def citation(record: dict[str, str]) -> str:
    if record.get("citation"):
        return record["citation"]
    parts = [record.get("authors", ""), record.get("year", ""), record.get("title", ""), record.get("journal", "")]
    return ". ".join(part for part in parts if part).strip()


def first_author(record: dict[str, str]) -> str:
    authors = text(record.get("authors"))
    if not authors:
        return "UnknownAuthor"
    author = re.split(r";|\band\b", authors, flags=re.I)[0].strip()
    if "," in author:
        author = author.split(",", 1)[0]
    else:
        author = author.split()[0] if author.split() else author
    return clean_component(author, "UnknownAuthor", 28)


def filename_for(record: dict[str, str], suffix: int = 1, filename_style: str = "legacy") -> str:
    author = first_author(record)
    year_match = re.search(r"(?:18|19|20)\d{2}", record.get("year", ""))
    year = year_match.group(0) if year_match else "n.d."
    if filename_style == "first-author-country-year":
        country = clean_label(record.get("first_author_country"), "UnknownCountry", 36)
        stem = f"{clean_label(author, 'UnknownAuthor', 28)} {country} {year}"
        return f"{stem}{'' if suffix == 1 else f' {suffix}'}.pdf"
    journal = clean_component(record.get("journal"), "Journal", 20)
    token = re.sub(r"[^a-z0-9]+", "", record.get("doi", "").lower())[-12:]
    token = token or ("PMID" + record["pmid"] if record["pmid"] else "noid")
    stem = f"{author}_{year}_{journal}_{token}"
    return f"{stem}{'' if suffix == 1 else f'_{suffix}'}.pdf"


def available_path(output_dir: Path, record: dict[str, str], filename_style: str = "legacy") -> Path:
    for number in range(1, 10000):
        candidate = output_dir / filename_for(record, number, filename_style)
        if not candidate.exists():
            return candidate
    raise RuntimeError("Could not allocate a unique PDF filename.")


def canonical_record(raw: dict[str, Any]) -> dict[str, str]:
    lower = {str(k).lower().replace(" ", "_"): v for k, v in raw.items()}

    def get(*names: str) -> str:
        for name in names:
            if lower.get(name) not in (None, ""):
                return text(lower[name])
        return ""

    return {
        "record_id": get("record_id", "id", "uid"),
        "pmid": normalize_pmid(get("pmid", "pubmed_id")),
        "doi": normalize_doi(get("doi", "doi_url")),
        "title": get("title", "article_title"),
        "authors": get("authors", "author", "first_author"),
        "year": get("year", "publication_year", "pub_year", "date"),
        "journal": get("journal", "source", "journal_title"),
        "citation": get("citation"),
        "first_author_country": get("first_author_country"),
    }


def load_json(path: Path) -> list[dict[str, str]]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(data, dict):
        data = data.get("records", data.get("results", []))
    if not isinstance(data, list):
        raise ValueError("JSON must contain a list or an object with a records list.")
    return [canonical_record(row) for row in data if isinstance(row, dict)]


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [canonical_record(row) for row in csv.DictReader(handle)]


def load_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX input needs openpyxl. Use the bundled workspace Python.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    fields = [text(value) for value in headers]
    return [
        canonical_record(dict(zip(fields, row)))
        for row in rows
        if any(value not in (None, "") for value in row)
    ]


def load_ris(path: Path) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    current: dict[str, str] = {}
    mapping = {
        "DO": "doi", "ID": "pmid", "TI": "title", "T1": "title",
        "AU": "authors", "PY": "year", "Y1": "year", "JO": "journal",
        "JF": "journal",
    }
    for line in path.read_text(encoding="utf-8-sig", errors="replace").splitlines():
        if len(line) >= 6 and line[2:6] == "  - ":
            tag, value = line[:2], line[6:].strip()
            if tag == "TY" and current:
                records.append(canonical_record(current))
                current = {}
            field = mapping.get(tag)
            if field:
                current[field] = (current.get(field, "") + "; " + value).strip("; ") if field == "authors" else value
        elif line.startswith("ER  -") and current:
            records.append(canonical_record(current))
            current = {}
    if current:
        records.append(canonical_record(current))
    return records


def load_input(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return load_json(path)
    if suffix == ".csv":
        return load_csv(path)
    if suffix == ".xlsx":
        return load_xlsx(path)
    if suffix == ".ris":
        return load_ris(path)
    raise ValueError("Supported inputs are JSON, CSV, XLSX, and RIS.")


def add_explicit_records(pmids: list[str], dois: list[str]) -> list[dict[str, str]]:
    empty = {"record_id": "", "title": "", "authors": "", "year": "", "journal": "", "citation": ""}
    records = [{**empty, "pmid": normalize_pmid(value), "doi": ""} for value in pmids]
    records += [{**empty, "pmid": "", "doi": normalize_doi(value)} for value in dois]
    return records


def dedupe(records: list[dict[str, str]]) -> list[dict[str, str]]:
    chosen: OrderedDict[str, dict[str, str]] = OrderedDict()
    for record in records:
        if not (record["pmid"] or record["doi"] or record["title"]):
            continue
        key = record_key(record)
        if key not in chosen:
            chosen[key] = record
        else:
            for field, value in record.items():
                if not chosen[key].get(field) and value:
                    chosen[key][field] = value
    return list(chosen.values())


def request_bytes(url: str, timeout: int = 45) -> tuple[bytes, str, str]:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": USER_AGENT,
            "Accept": "application/pdf, application/json;q=0.9, */*;q=0.1",
        },
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read(), response.headers.get("Content-Type", ""), response.geturl()


def request_json(url: str) -> dict[str, Any] | None:
    try:
        data, _, _ = request_bytes(url)
        return json.loads(data.decode("utf-8"))
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, ValueError, UnicodeDecodeError):
        return None


def pubmed_metadata(pmid: str) -> dict[str, str]:
    if not pmid:
        return {}
    email = os.environ.get("NCBI_EMAIL", "researcher@example.com")
    params = urllib.parse.urlencode(
        {
            "db": "pubmed", "id": pmid, "retmode": "json",
            "tool": "retrieve-journal-pdfs", "email": email,
        }
    )
    payload = request_json(EUTILS + "esummary.fcgi?" + params) or {}
    article = payload.get("result", {}).get(pmid, {})
    ids = {
        text(item.get("idtype")).lower(): text(item.get("value"))
        for item in article.get("articleids", [])
        if isinstance(item, dict)
    }
    authors = "; ".join(
        text(author.get("name"))
        for author in article.get("authors", [])[:6]
        if isinstance(author, dict)
    )
    return {
        "pmcid": ids.get("pmc", "").upper(),
        "doi": normalize_doi(ids.get("doi", "")),
        "title": text(article.get("title")).rstrip("."),
        "authors": authors,
        "year": text(article.get("pubdate"))[:4],
        "journal": text(article.get("source")),
    }


COUNTRY_ALIASES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("Australia", ("australia",)),
    ("Austria", ("austria",)),
    ("Belgium", ("belgium",)),
    ("Brazil", ("brazil",)),
    ("Canada", ("canada",)),
    ("China", ("china",)),
    ("Czech Republic", ("czech republic", "czechia")),
    ("Denmark", ("denmark",)),
    ("Finland", ("finland",)),
    ("France", ("france",)),
    ("Germany", ("germany",)),
    ("Greece", ("greece",)),
    ("Hong Kong", ("hong kong",)),
    ("India", ("india",)),
    ("Ireland", ("ireland",)),
    ("Israel", ("israel",)),
    ("Italy", ("italy",)),
    ("Japan", ("japan",)),
    ("Korea", ("south korea", "republic of korea", "korea")),
    ("Malaysia", ("malaysia",)),
    ("Mexico", ("mexico",)),
    ("Netherlands", ("netherlands", "the netherlands")),
    ("New Zealand", ("new zealand",)),
    ("Norway", ("norway",)),
    ("Poland", ("poland",)),
    ("Portugal", ("portugal",)),
    ("Saudi Arabia", ("saudi arabia",)),
    ("Singapore", ("singapore",)),
    ("South Africa", ("south africa",)),
    ("Spain", ("spain",)),
    ("Sweden", ("sweden",)),
    ("Switzerland", ("switzerland",)),
    ("Taiwan", ("taiwan",)),
    ("Thailand", ("thailand",)),
    ("Turkey", ("turkey", "türkiye")),
    ("United Kingdom", ("united kingdom", "uk", "u.k.")),
    ("United States", ("united states", "usa", "u.s.a.")),
)


def country_from_affiliation(value: str) -> str:
    normalized = " " + re.sub(r"[^a-z]+", " ", value.casefold()) + " "
    for country, aliases in COUNTRY_ALIASES:
        for alias in aliases:
            alias_words = " " + re.sub(r"[^a-z]+", " ", alias.casefold()) + " "
            if alias_words in normalized:
                return country
    return ""


def pubmed_first_author_country(pmid: str) -> str:
    if not pmid:
        return ""
    email = os.environ.get("NCBI_EMAIL", "researcher@example.com")
    params = urllib.parse.urlencode(
        {
            "db": "pubmed", "id": pmid, "retmode": "xml",
            "tool": "retrieve-journal-pdfs", "email": email,
        }
    )
    try:
        data, _, _ = request_bytes(EUTILS + "efetch.fcgi?" + params)
        root = ET.fromstring(data)
    except (ET.ParseError, urllib.error.URLError, urllib.error.HTTPError, TimeoutError):
        return ""
    author = root.find(".//AuthorList/Author")
    if author is None:
        return ""
    affiliations = [
        "".join(node.itertext()).strip()
        for node in author.findall("./AffiliationInfo/Affiliation")
    ]
    for affiliation in affiliations:
        country = country_from_affiliation(affiliation)
        if country:
            return country
    return ""


def europe_pmc_metadata(record: dict[str, str]) -> dict[str, str]:
    query = f"EXT_ID:{record['pmid']}" if record["pmid"] else f'DOI:"{record["doi"]}"'
    url = EUROPE_PMC + "?" + urllib.parse.urlencode({"query": query, "format": "json", "pageSize": "1"})
    payload = request_json(url) or {}
    results = payload.get("resultList", {}).get("result", [])
    if not results:
        return {}
    result = results[0]
    return {
        "pmcid": text(result.get("pmcid")).upper(),
        "doi": normalize_doi(result.get("doi")),
        "title": text(result.get("title")),
        "authors": text(result.get("authorString")),
        "year": text(result.get("pubYear")),
        "journal": text(result.get("journalTitle")),
    }


def merge_metadata(record: dict[str, str], *metadata: dict[str, str]) -> dict[str, str]:
    merged = dict(record)
    for meta in metadata:
        for field in ("doi", "title", "authors", "year", "journal"):
            if not merged.get(field) and meta.get(field):
                merged[field] = meta[field]
    return merged


def candidate_urls(record: dict[str, str], pubmed: dict[str, str], europe: dict[str, str]) -> list[dict[str, str]]:
    candidates: list[dict[str, str]] = []
    pmcid = pubmed.get("pmcid") or europe.get("pmcid")
    if pmcid:
        candidates.append({"source": "PMC", "access_type": "open", "url": f"https://pmc.ncbi.nlm.nih.gov/articles/{pmcid}/pdf/"})
        candidates.append({"source": "Europe PMC", "access_type": "open", "url": f"https://europepmc.org/articles/{pmcid}?pdf=render"})
    email = os.environ.get("UNPAYWALL_EMAIL") or os.environ.get("NCBI_EMAIL")
    if record["doi"] and email:
        url = UNPAYWALL + urllib.parse.quote(record["doi"], safe="") + "?" + urllib.parse.urlencode({"email": email})
        payload = request_json(url) or {}
        locations = [payload.get("best_oa_location")] + list(payload.get("oa_locations") or [])
        seen: set[str] = set()
        for location in locations:
            if not isinstance(location, dict):
                continue
            candidate = text(location.get("url_for_pdf"))
            if candidate and candidate not in seen:
                seen.add(candidate)
                candidates.append({"source": text(location.get("host_type")) or "Unpaywall", "access_type": "open", "url": candidate})
    if record["doi"]:
        candidates.append({"source": "DOI publisher landing page", "access_type": "browser_handoff", "url": "https://doi.org/" + record["doi"]})
    return candidates


def validate_pdf(data: bytes) -> tuple[bool, str, str]:
    if not data.startswith(b"%PDF-"):
        return False, "", "Response did not begin with a PDF header."
    try:
        from pypdf import PdfReader
        return True, str(len(PdfReader(io.BytesIO(data)).pages)), ""
    except ImportError:
        if b"%%EOF" not in data[-4096:]:
            return False, "", "PDF EOF marker was absent."
    except Exception as exc:
        return False, "", f"PDF parser rejected the file: {type(exc).__name__}."
    return True, "", ""


def now_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def initial_result(record: dict[str, str], filename_style: str = "legacy") -> dict[str, str]:
    return {
        "record_id": record["record_id"] or record_key(record),
        "citation": citation(record), "pmid": record["pmid"], "doi": record["doi"],
        "title": record["title"], "retrieval_status": "", "source": "",
        "access_type": "", "candidate_url": "", "local_path_or_url": "",
        "content_type": "", "sha256": "", "page_count": "", "notes": "",
        "next_action": "", "retrieved_utc": now_utc(),
        "first_author_country": record.get("first_author_country", ""),
        "filename_style": filename_style,
    }


def country_note(record: dict[str, str], filename_style: str, message: str = "") -> str:
    if filename_style != "first-author-country-year":
        return message
    country_message = (
        "First author country verified from PubMed affiliation."
        if record.get("first_author_country")
        else "First author country could not be verified from PubMed affiliation; filename uses UnknownCountry."
    )
    return f"{country_message} {message}".strip()


def retrieve(
    record: dict[str, str],
    output_dir: Path,
    dry_run: bool,
    browser_session: bool,
    filename_style: str = "legacy",
) -> dict[str, str]:
    if not (record["pmid"] or record["doi"]):
        result = initial_result(record, filename_style)
        result.update(
            retrieval_status="input_error",
            notes="Record needs a PMID or DOI; title-only matching is not automated.",
            next_action="Add a verified PMID or DOI.",
        )
        return result
    pubmed = pubmed_metadata(record["pmid"])
    if record["doi"] and pubmed.get("doi") and record["doi"] != pubmed["doi"]:
        result = initial_result(record, filename_style)
        result.update(
            retrieval_status="input_error",
            notes=f"Input DOI conflicts with PubMed PMID {record['pmid']} metadata.",
            next_action="Correct the PMID or DOI before retrieving full text.",
        )
        return result
    europe = europe_pmc_metadata(record)
    record = merge_metadata(record, pubmed, europe)
    if filename_style == "first-author-country-year" and not record.get("first_author_country"):
        record["first_author_country"] = pubmed_first_author_country(record["pmid"])
    result = initial_result(record, filename_style)
    candidates = candidate_urls(record, pubmed, europe)
    open_candidates = [item for item in candidates if item["access_type"] == "open"]
    if dry_run and open_candidates:
        item = open_candidates[0]
        result.update(
            retrieval_status="candidate_found_dry_run", source=item["source"],
            access_type="open", candidate_url=item["url"],
            local_path_or_url=item["url"],
            notes=country_note(record, filename_style),
            next_action="Re-run without --dry-run to download and validate.",
        )
        return result
    for item in open_candidates:
        try:
            data, content_type, final_url = request_bytes(item["url"])
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as exc:
            result.update(
                retrieval_status="download_failed", source=item["source"],
                access_type="open", candidate_url=item["url"],
                notes=country_note(record, filename_style, f"{type(exc).__name__}: {exc}"),
                next_action="Review candidate URL or retry later.",
            )
            continue
        valid, pages, note = validate_pdf(data)
        if not valid:
            result.update(
                retrieval_status="invalid_pdf", source=item["source"],
                access_type="open", candidate_url=final_url,
                content_type=content_type, notes=country_note(record, filename_style, note),
                next_action="Inspect the source manually; do not treat this response as a PDF.",
            )
            continue
        output_dir.mkdir(parents=True, exist_ok=True)
        destination = available_path(output_dir, record, filename_style)
        destination.write_bytes(data)
        result.update(
            retrieval_status="retrieved", source=item["source"], access_type="open",
            candidate_url=final_url, local_path_or_url=str(destination.resolve()),
            content_type=content_type, sha256=hashlib.sha256(data).hexdigest(),
            page_count=pages, notes=country_note(record, filename_style, "Verified PDF."),
            next_action="Ready for full-text screening.",
        )
        return result
    publisher = next((item for item in candidates if item["access_type"] == "browser_handoff"), None)
    if publisher:
        result.update(
            retrieval_status="needs_browser_session" if browser_session else "not_found",
            source=publisher["source"], access_type=publisher["access_type"],
            candidate_url=publisher["url"], local_path_or_url=publisher["url"],
            notes=country_note(record, filename_style, "No supported open-access PDF was located."),
            next_action=(
                "Open this page only in an already authenticated browser session."
                if browser_session
                else "Check library access or rerun with --browser-session for a safe handoff link."
            ),
        )
    elif not result["retrieval_status"]:
        result.update(
            retrieval_status="not_found",
            notes=country_note(record, filename_style, "No supported legal full-text source was located."),
            next_action="Check library holdings, author manuscript, or document delivery.",
        )
    return result


def read_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if not path.exists():
        return [], []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), [dict(row) for row in reader]


def row_key(row: dict[str, str]) -> str:
    doi = normalize_doi(row.get("doi"))
    pmid = normalize_pmid(row.get("pmid"))
    return "doi:" + doi if doi else ("pmid:" + pmid if pmid else text(row.get("record_id")))


def merge_manifest(path: Path, headers: list[str], new_rows: list[dict[str, str]], preserve_retrieved: bool = True) -> None:
    old_headers, old_rows = read_rows(path)
    final_headers = list(dict.fromkeys(old_headers + headers))
    indexed = OrderedDict((row_key(row), row) for row in old_rows if row_key(row))
    for new_row in new_rows:
        key = row_key(new_row)
        old = indexed.get(key)
        if old and preserve_retrieved and text(old.get("retrieval_status")).lower() == "retrieved" and text(new_row.get("retrieval_status")).lower() != "retrieved":
            continue
        merged = dict(old or {})
        merged.update({field: value for field, value in new_row.items() if value != ""})
        indexed[key] = merged
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=final_headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(indexed.values())


def project_root_for(input_path: Path | None, output_arg: str | None) -> Path | None:
    starts: list[Path] = []
    if input_path:
        starts.append(input_path.parent)
    if output_arg:
        starts.append(Path(output_arg).resolve())
    for start in starts:
        for parent in [start, *start.parents]:
            if (parent / "04_fulltext").is_dir() and (parent / "03_screening").is_dir():
                return parent
    return None


def parse_many(values: list[str] | None) -> list[str]:
    return [item.strip() for value in (values or []) for item in value.split(",") if item.strip()]


def main() -> int:
    parser = argparse.ArgumentParser(description="Find and retrieve legally accessible journal PDFs.")
    parser.add_argument("--input", help="JSON/CSV/XLSX/RIS records, including narrative-review-search records.json.")
    parser.add_argument("--pmid", action="append", help="PMID or comma-separated PMIDs; repeatable.")
    parser.add_argument("--doi", action="append", help="DOI or comma-separated DOIs; repeatable.")
    parser.add_argument("--output-dir", help="PDF destination; default is fulltext/ or SRMA 04_fulltext/.")
    parser.add_argument(
        "--filename-style",
        choices=("legacy", "first-author-country-year"),
        default="legacy",
        help="PDF filename format; default keeps the existing author_year_journal_identifier format.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Find legal candidates without downloading files.")
    parser.add_argument("--browser-session", action="store_true", help="Emit a safe publisher handoff for records without open PDFs; never reads cookies or logs in.")
    args = parser.parse_args()
    if not (args.input or args.pmid or args.doi):
        parser.error("Provide --input, --pmid, or --doi.")
    input_path = Path(args.input).resolve() if args.input else None
    records: list[dict[str, str]] = load_input(input_path) if input_path else []
    records += add_explicit_records(parse_many(args.pmid), parse_many(args.doi))
    records = dedupe(records)
    if not records:
        print("No usable records found.", file=sys.stderr)
        return 2
    project_root = project_root_for(input_path, args.output_dir)
    output_dir = (
        Path(args.output_dir).resolve()
        if args.output_dir
        else (project_root / "04_fulltext" if project_root else Path.cwd() / "fulltext")
    )
    results: list[dict[str, str]] = []
    for number, record in enumerate(records, start=1):
        print(f"[{number}/{len(records)}] PMID={record['pmid'] or '-'} DOI={record['doi'] or '-'}")
        results.append(
            retrieve(record, output_dir, args.dry_run, args.browser_session, args.filename_style)
        )
        time.sleep(0.35)
    merge_manifest(output_dir / "retrieval_manifest.csv", DETAIL_HEADERS, results)
    if project_root or output_dir.name.lower() == "04_fulltext":
        srma_rows = [{header: result.get(header, "") for header in SRMA_HEADERS} for result in results]
        merge_manifest(output_dir / "manifest.csv", SRMA_HEADERS, srma_rows)
    summary: dict[str, int] = {}
    for result in results:
        summary[result["retrieval_status"]] = summary.get(result["retrieval_status"], 0) + 1
    print("Output:", output_dir)
    print("Summary:", ", ".join(f"{status}={count}" for status, count in sorted(summary.items())))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
