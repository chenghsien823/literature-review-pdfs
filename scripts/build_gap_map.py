#!/usr/bin/env python3
"""
build_gap_map.py — Turn an evidence-extractor study list into an evidence gap map
(contrast x outcome-domain matrix) + a gap-analysis worksheet. Mechanical density
only; the gap STATEMENT and proposed study stay blank for a human.

Usage:
    python3 build_gap_map.py studies.json gap_map.xlsx

Input JSON: the evidence-extractor study list. Each study uses shared columns
(design, evidence_level, direction, conflict, citation) plus, for the map axes:
    contrast : str   (e.g. "benralizumab vs mepolizumab")   [top-level or in fields]
    domains  : [str] (e.g. ["remission","GC-sparing"])       [top-level or in fields]
Missing contrast/domains -> study counted as 'unclassified' and flagged.
"""
import sys, json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LEVEL_FILL = {"High": PatternFill("solid", fgColor="C6EFCE"),
              "Moderate": PatternFill("solid", fgColor="FFEB9C"),
              "Low": PatternFill("solid", fgColor="FFC7CE")}
EMPTY_FILL = PatternFill("solid", fgColor="F2F2F2")        # candidate gap
GAP_FILL = PatternFill("solid", fgColor="FFF2CC")          # human-decision cols
CONFLICT_FILL = PatternFill("solid", fgColor="FFD9D9")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

DESIGN_RANK = {"Meta-analysis": 6, "Systematic review": 5, "RCT": 4,
               "Cohort": 3, "Case-control": 3, "Cross-sectional": 2,
               "Case series": 1, "Case report": 1}
LEVEL_RANK = {"High": 3, "Moderate": 2, "Low": 1}
DIR_SYM = {"support": "↑", "oppose": "↓", "neutral": "→"}

def get(study, key, default=None):
    if key in study and study[key] not in (None, ""):
        return study[key]
    f = study.get("fields", {}) or {}
    return f.get(key, default)

def main():
    if len(sys.argv) != 3:
        print(__doc__); sys.exit(1)
    studies = json.load(open(sys.argv[1], encoding="utf-8-sig"))
    out_path = sys.argv[2]

    cells = defaultdict(list)        # (contrast, domain) -> [study,...]
    contrasts, domains = [], []
    unclassified = []
    for st in studies:
        contrast = get(st, "contrast")
        doms = get(st, "domains") or []
        if not contrast or not doms:
            unclassified.append(st.get("citation", "?"))
            if not contrast: contrast = "unclassified"
            if not doms: doms = ["unclassified"]
        if contrast not in contrasts: contrasts.append(contrast)
        for d in doms:
            if d not in domains: domains.append(d)
            cells[(contrast, d)].append(st)

    def cell_summary(studs):
        if not studs:
            return None
        n = len(studs)
        top = max(studs, key=lambda s: DESIGN_RANK.get(s.get("design",""), 0))
        top_design = top.get("design", "?")
        dirs = {get(s, "direction", "neutral") for s in studs}
        net = "mixed" if len({d for d in dirs if d in ("support","oppose")}) > 1 else \
              (DIR_SYM.get(next(iter(dirs)), "→") if dirs else "→")
        if net == "mixed": net = "↕ mixed"
        conflict = any((s.get("conflict") or "").strip() for s in studs)
        best_level = max((s.get("evidence_level","Low") for s in studs),
                         key=lambda l: LEVEL_RANK.get(l,1))
        return {"n": n, "top_design": top_design, "net": net,
                "conflict": conflict, "level": best_level}

    wb = Workbook()

    # ---- Gap Map ----
    ws = wb.active; ws.title = "Gap Map"
    ws.cell(1,1,"contrast \\ domain").fill = HEADER_FILL
    ws.cell(1,1).font = HEADER_FONT; ws.cell(1,1).alignment = CENTER; ws.cell(1,1).border = BORDER
    for c,d in enumerate(domains, start=2):
        cell = ws.cell(1,c,d); cell.fill=HEADER_FILL; cell.font=HEADER_FONT
        cell.alignment=CENTER; cell.border=BORDER
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.column_dimensions["A"].width = 26
    for r,contrast in enumerate(contrasts, start=2):
        rc = ws.cell(r,1,contrast); rc.font=Font(bold=True); rc.alignment=WRAP; rc.border=BORDER
        for c,d in enumerate(domains, start=2):
            summ = cell_summary(cells.get((contrast,d), []))
            cell = ws.cell(r,c); cell.alignment=CENTER; cell.border=BORDER
            if summ is None:
                cell.value = "—  GAP?"; cell.fill = EMPTY_FILL
            else:
                flag = " ⚠" if summ["conflict"] else ""
                cell.value = f"{summ['n']} · {summ['top_design']} · {summ['net']}{flag}"
                cell.fill = CONFLICT_FILL if summ["conflict"] else LEVEL_FILL.get(summ["level"], EMPTY_FILL)
    ws.freeze_panes = "B2"

    # ---- Gap Worksheet ----
    gw = wb.create_sheet("Gap Worksheet")
    cols = [("contrast",22),("domain",16),("n_studies",9),("top_design",14),
            ("net_direction",13),("conflict",10),("best_evidence",13),("studies",30),
            ("GAP_STATEMENT",40),("PROPOSED_STUDY",34)]
    for c,(name,w) in enumerate(cols,1):
        cell=gw.cell(1,c,name); cell.fill=HEADER_FILL; cell.font=HEADER_FONT
        cell.alignment=CENTER; cell.border=BORDER
        gw.column_dimensions[get_column_letter(c)].width=w
    row=2
    # every contrast x domain pair, INCLUDING empty (the candidate gaps)
    for contrast in contrasts:
        for d in domains:
            studs=cells.get((contrast,d),[])
            summ=cell_summary(studs)
            cites="; ".join(s.get("citation","?") for s in studs)
            vals=[contrast,d,
                  summ["n"] if summ else 0,
                  summ["top_design"] if summ else "—",
                  summ["net"] if summ else "—",
                  "yes" if (summ and summ["conflict"]) else "",
                  summ["level"] if summ else "EMPTY",
                  cites,"",""]
            for c,v in enumerate(vals,1):
                cell=gw.cell(row,c,v); cell.alignment=WRAP; cell.border=BORDER
                if c in (9,10): cell.fill=GAP_FILL
                if not summ and c==7: cell.fill=EMPTY_FILL
            row+=1
    gw.freeze_panes="A2"
    gw.auto_filter.ref=f"A1:{get_column_letter(len(cols))}{row-1}"

    # ---- Coverage Log ----
    log=wb.create_sheet("Coverage Log")
    lines=[["Research-gap coverage log",""],["",""],
           ["Studies counted:",len(studies)],
           ["Unclassified (missing contrast/domains):",len(unclassified)],
           ["  -> "+", ".join(unclassified) if unclassified else "  -> none",""],
           ["Contrasts seen:",len([c for c in contrasts if c!='unclassified'])],
           ["Domains seen:",len([d for d in domains if d!='unclassified'])],
           ["",""],
           ["NOTE: an empty cell is a gap ONLY within these extracted studies.",""],
           ["The gap statement + its search-scope caveat are the reviewer's to write.",""]]
    for r,(a,b) in enumerate(lines,1):
        ca=log.cell(r,1,a); log.cell(r,2,b)
        if r==1: ca.font=Font(bold=True,size=13)
        elif a.endswith(":"): ca.font=Font(bold=True)
    log.column_dimensions["A"].width=58; log.column_dimensions["B"].width=12

    wb.save(out_path)
    gaps=sum(1 for contrast in contrasts for d in domains if not cells.get((contrast,d)))
    print(f"Wrote {out_path}: {len(contrasts)} contrasts x {len(domains)} domains | "
          f"{gaps} empty cells (candidate gaps) | {len(unclassified)} unclassified")

if __name__ == "__main__":
    main()
