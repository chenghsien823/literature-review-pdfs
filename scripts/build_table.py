#!/usr/bin/env python3
"""
build_table.py — Build a formatted evidence table (xlsx) from extracted study data.

Usage:
    python3 build_table.py input.json output.xlsx

Input JSON: a list of study objects. Each object has the SHARED columns plus a
`fields` object holding the design-specific schema (see references/*.md).

Shared keys (all studies):
    citation, pmid, doi, verified (bool), design, population, key_finding,
    evidence_level ("High"|"Moderate"|"Low"),
    direction ("support"|"oppose"|"neutral"),
    conflict (str, "" if none), synthesis_note (str)
Design-specific keys live under: fields = { ... }  (see reference schemas)

Output: xlsx with
  - "Synthesis Master" sheet: shared columns, colour-coded evidence level,
    direction symbols, conflict highlighting, verification flag, frozen panes, autofilter.
  - One detail sheet per design present: shared id columns + that design's full fields.
"""
import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- styling ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
LEVEL_FILL = {"High": PatternFill("solid", fgColor="C6EFCE"),
              "Moderate": PatternFill("solid", fgColor="FFEB9C"),
              "Low": PatternFill("solid", fgColor="FFC7CE")}
CONFLICT_FILL = PatternFill("solid", fgColor="FFD9D9")
UNVERIFIED_FILL = PatternFill("solid", fgColor="FF7F7F")
DIRECTION = {"support": "↑ support", "oppose": "↓ oppose", "neutral": "→ neutral"}
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

MASTER_COLS = [
    ("citation", "Citation", 16), ("design", "Design", 14),
    ("verified", "Verified", 9), ("pmid", "PMID", 12), ("doi", "DOI", 18),
    ("population", "Population", 22), ("key_finding", "Key finding", 34),
    ("evidence_level", "Evidence", 11), ("direction", "Direction", 12),
    ("conflict", "Conflict?", 20), ("synthesis_note", "Synthesis note", 30),
]

# design -> ordered (key, label, width) from reference schemas
DESIGN_FIELDS = {
    "RCT": [("inclusion","Inclusion",22),("intervention","Intervention",20),("comparator","Comparator",18),
            ("randomization","Randomization/concealment",24),("blinding","Blinding",16),("n_total","N (per arm)",14),
            ("primary_outcome","Primary outcome",22),("effect","Effect + 95% CI",22),("analysis","Analysis",12),
            ("followup","Follow-up",14),("funding_coi","Funding/COI",18)],
    "Cohort": [("exposure","Exposure",18),("comparator","Comparator",16),("outcome","Outcome",18),
               ("confounding","Confounding adjustment",26),("effect","Effect + 95% CI",22),("followup","Follow-up",14),
               ("incidence","Incidence",16),("n_total","N",10),("source","Data source",18)],
    "Case-control": [("case_def","Case definition",22),("control_source","Controls + matching",24),
                     ("exposure","Exposure",18),("exposure_ascertain","Exposure ascertainment",24),
                     ("confounding","Confounding adjustment",26),("effect","OR + 95% CI",20),("n_total","N",10),("source","Source",16)],
    "Cross-sectional": [("sampling","Sampling",20),("exposure","Exposure",18),("outcome","Outcome",18),
                        ("confounding","Adjustment",22),("effect","Assoc. + 95% CI",22),("temporality","Temporality caveat",24),
                        ("n_total","N",10),("source","Source",16)],
    "Case report": [("n","N",8),("patient","Patient(s)",26),("intervention","Intervention",20),
                    ("outcome","Outcome",22),("timeline","Timeline",20),("generalizability","Generalizability caveat",28)],
    "Case series": [("n","N",8),("patient","Patient(s)",26),("intervention","Intervention",20),
                    ("outcome","Outcome",22),("timeline","Timeline",20),("generalizability","Generalizability caveat",28)],
    "Systematic review": [("question","Question (PICO)",26),("databases","Databases + date",22),("n_studies","N studies",12),
                          ("pooled_effect","Pooled effect + 95% CI",24),("heterogeneity","Heterogeneity (I²)",16),
                          ("quality_tool","RoB/quality tool",18),("included_designs","Included designs",20)],
    "Meta-analysis": [("question","Question (PICO)",26),("databases","Databases + date",22),("n_studies","N studies",12),
                      ("pooled_effect","Pooled effect + 95% CI",24),("heterogeneity","Heterogeneity (I²)",16),
                      ("quality_tool","RoB/quality tool",18),("included_designs","Included designs",20)],
}

def style_header(ws, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(1, c); cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 30

def build_master(wb, studies):
    ws = wb.active; ws.title = "Synthesis Master"
    for i,(k,label,w) in enumerate(MASTER_COLS, start=1):
        ws.cell(1,i,label); ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, len(MASTER_COLS))
    for r, s in enumerate(studies, start=2):
        for i,(k,label,w) in enumerate(MASTER_COLS, start=1):
            v = s.get(k, "")
            if k == "verified": v = "✓" if s.get("verified") else "✗ UNVERIFIED"
            if k == "direction": v = DIRECTION.get(s.get("direction",""), s.get("direction",""))
            cell = ws.cell(r,i,v); cell.alignment = WRAP; cell.border = BORDER
            if k == "evidence_level" and v in LEVEL_FILL:
                cell.fill = LEVEL_FILL[v]; cell.alignment = Alignment(horizontal="center", vertical="top")
            if k == "conflict" and str(v).strip():
                cell.fill = CONFLICT_FILL
            if k == "verified" and not s.get("verified"):
                cell.fill = UNVERIFIED_FILL; cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="top")
        ws.row_dimensions[r].height = 42
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLS))}{len(studies)+1}"
    return ws

def build_detail(wb, design, studies):
    fields = DESIGN_FIELDS[design]
    cols = [("citation","Citation",16),("pmid","PMID",12),("doi","DOI",18)] + fields
    ws = wb.create_sheet(design[:31])
    for i,(k,label,w) in enumerate(cols, start=1):
        ws.cell(1,i,label); ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, len(cols))
    r = 2
    for s in studies:
        if s.get("design") != design: continue
        f = s.get("fields", {})
        for i,(k,label,w) in enumerate(cols, start=1):
            v = s.get(k, "") if k in ("citation","pmid","doi") else f.get(k, "NR")
            cell = ws.cell(r,i,v); cell.alignment = WRAP; cell.border = BORDER
        ws.row_dimensions[r].height = 42; r += 1
    ws.freeze_panes = "B2"
    if r > 2: ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{r-1}"

def main():
    if len(sys.argv) != 3:
        print("Usage: python3 build_table.py input.json output.xlsx"); sys.exit(1)
    studies = json.load(open(sys.argv[1], encoding="utf-8-sig"))
    wb = Workbook()
    build_master(wb, studies)
    for design in DESIGN_FIELDS:
        if any(s.get("design") == design for s in studies):
            build_detail(wb, design, studies)
    wb.save(sys.argv[2])
    print(f"Wrote {sys.argv[2]}: {len(studies)} studies, "
          f"{1+sum(1 for d in DESIGN_FIELDS if any(s.get('design')==d for s in studies))} sheets")

if __name__ == "__main__":
    main()
