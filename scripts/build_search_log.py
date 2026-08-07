#!/usr/bin/env python3
"""
build_search_log.py — Export a raw narrative-review-search sweep (the full set of
records identified) into an auditable Excel SEARCH LOG. This is the identification
pool BEFORE de-duplication/screening — the search-record artefact a review keeps
for transparency / PRISMA "records identified".

Usage:
    python3 build_search_log.py records.json search_log.xlsx
        [--query "..."] [--count N] [--sort relevance] [--label "EGPA scoping sweep"]

Provenance: if a sibling "<records>.meta.json" exists (written automatically by
search.py), its query/count/sort/database/retrieved_utc are used. CLI flags override.
"""
import sys, os, json, argparse, time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
INFO_KEY = Font(bold=True)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

COLS = [("rank",6),("pmid",11),("doi",22),("year",7),("journal",20),
        ("pub_type",24),("title",54),("authors",30),("abstract",60),("mesh",30),("url",30)]

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("records"); ap.add_argument("out")
    ap.add_argument("--query"); ap.add_argument("--count")
    ap.add_argument("--sort"); ap.add_argument("--label")
    ap.add_argument("--database", default="PubMed (E-utilities)")
    a = ap.parse_args()

    recs = json.load(open(a.records, encoding="utf-8-sig"))
    meta = {}
    sidecar = a.records + ".meta.json"
    if os.path.exists(sidecar):
        meta = json.load(open(sidecar, encoding="utf-8-sig"))
    query = a.query or meta.get("query", "NR")
    count = a.count or meta.get("count", "NR")
    sort = a.sort or meta.get("sort", "NR")
    database = a.database or meta.get("database", "PubMed")
    retrieved = meta.get("retrieved_utc", time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime()))

    wb = Workbook()

    # ---- Search Info ----
    info = wb.active; info.title = "Search Info"
    rows = [
        ("Search log — records identified (raw sweep)", ""),
        ("", ""),
        ("Label", a.label or ""),
        ("Database", database),
        ("Query", query),
        ("Sort", sort),
        ("Total hits (count)", count),
        ("Records in this file", len(recs)),
        ("Retrieved (UTC)", retrieved),
        ("", ""),
        ("NOTE", "This is the identification pool BEFORE de-duplication and screening."),
        ("", "De-dup + include/exclude happen in lit-screen; this file is the audit trail."),
        ("", "If 'Records in this file' < 'Total hits', raise max_results to capture the full pool."),
    ]
    for r,(k,v) in enumerate(rows,1):
        ck=info.cell(r,1,k); info.cell(r,2,v)
        if r==1: ck.font=Font(bold=True,size=13)
        elif k and k!="NOTE": ck.font=INFO_KEY
        elif k=="NOTE": ck.font=INFO_KEY
    info.column_dimensions["A"].width=22; info.column_dimensions["B"].width=90

    # ---- Records ----
    ws = wb.create_sheet("Records")
    for c,(name,w) in enumerate(COLS,1):
        cell=ws.cell(1,c,name); cell.fill=HEADER_FILL; cell.font=HEADER_FONT
        cell.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center"); cell.border=BORDER
        ws.column_dimensions[get_column_letter(c)].width=w
    for i,r in enumerate(recs, start=1):
        pmid=r.get("pmid","")
        url=f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/" if pmid else (
            f"https://doi.org/{r.get('doi')}" if r.get("doi") not in ("","NR",None) else "")
        row=[i, pmid, r.get("doi",""), r.get("year",""), r.get("journal",""),
             r.get("publication_type",""), r.get("title",""), r.get("authors",""),
             r.get("abstract",""), r.get("mesh",""), url]
        for c,v in enumerate(row,1):
            cell=ws.cell(i+1,c,v); cell.alignment=WRAP; cell.border=BORDER
    ws.freeze_panes="C2"
    ws.auto_filter.ref=f"A1:{get_column_letter(len(COLS))}{len(recs)+1}"

    wb.save(a.out)
    print(f"Wrote {a.out}: {len(recs)} records (total hits={count}) | sort={sort}")

if __name__ == "__main__":
    main()
