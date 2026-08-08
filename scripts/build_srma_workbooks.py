#!/usr/bin/env python3
"""Create SRMA screening, extraction, RoB, and meta-analysis input workbooks."""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def load_records(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(payload, dict):
        payload = payload.get("records", payload.get("studies", []))
    if not isinstance(payload, list) or not all(isinstance(row, dict) for row in payload):
        raise ValueError("records_json must contain a list of record objects.")
    return payload


def study_id(record: dict[str, Any]) -> str:
    if str(record.get("pmid") or "").strip():
        return f"PMID:{record['pmid']}"
    if str(record.get("doi") or "").strip():
        return f"DOI:{record['doi']}"
    return f"TITLE:{str(record.get('title') or '').strip()[:80]}"


def citation(record: dict[str, Any]) -> str:
    author = str(record.get("authors") or "").split(";")[0].strip()
    return f"{author} {record.get('year') or ''}".strip()


def add_sheet(wb: Workbook, title: str, columns: list[tuple[str, int]], rows: list[list[Any]] | None = None) -> None:
    ws = wb.create_sheet(title[:31])
    for index, (label, width) in enumerate(columns, 1):
        cell = ws.cell(1, index, label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.row_dimensions[1].height = 36
    for row_index, row in enumerate(rows or [], 2):
        for column_index, value in enumerate(row, 1):
            cell = ws.cell(row_index, column_index, value)
            cell.alignment = WRAP
            cell.border = BORDER
        ws.row_dimensions[row_index].height = 42
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(rows or []) + 1)}"


def add_note_sheet(wb: Workbook, title: str, rows: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet(title[:31])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100
    for index, (key, value) in enumerate(rows, 1):
        a, b = ws.cell(index, 1, key), ws.cell(index, 2, value)
        a.font = Font(bold=True)
        a.fill = NOTE_FILL
        a.alignment = WRAP
        b.alignment = WRAP
    ws.freeze_panes = "A2"


def screening_workbook(records: list[dict[str, Any]], outdir: Path, query: dict[str, Any]) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    add_note_sheet(wb, "Read me", [
        ("Review type", "SRMA"),
        ("Authoritative decisions", "Keep canonical AI and human decisions in 02_screening JSONL. This workbook is an auditable review snapshot, not the source of truth."),
        ("Required gate", "Complete independent human title/abstract and full-text screening, then resolve conflicts before extraction."),
        ("Review question", str(query.get("review_question") or query.get("population") or "Define in protocol")),
    ])
    source_rows = [[study_id(r), citation(r), r.get("pmid", ""), r.get("doi", ""), r.get("year", ""), r.get("title", ""), r.get("abstract", ""), "", "", "", "", ""] for r in records]
    add_sheet(wb, "Title abstract", [
        ("Study ID", 20), ("Citation", 20), ("PMID", 13), ("DOI", 24), ("Year", 9), ("Title", 52), ("Abstract", 70),
        ("Reviewer 1 decision", 18), ("Reviewer 2 decision", 18), ("Exclusion reason", 26), ("Conflict resolution", 22), ("Decision date", 16),
    ], source_rows)
    fulltext_rows = [[study_id(r), citation(r), r.get("pmid", ""), r.get("doi", ""), r.get("title", ""), "", "", "", "", ""] for r in records]
    add_sheet(wb, "Full text", [
        ("Study ID", 20), ("Citation", 20), ("PMID", 13), ("DOI", 24), ("Title", 52),
        ("Reviewer 1 decision", 18), ("Reviewer 2 decision", 18), ("Exclusion reason", 26), ("Conflict resolution", 22), ("Decision date", 16),
    ], fulltext_rows)
    add_sheet(wb, "PRISMA counts", [("Flow stage", 36), ("Count", 16), ("Notes", 70)], [
        ["Records identified", len(records), "From 01_search_log.xlsx / records.json"],
        ["Records after deduplication", "", ""], ["Title/abstract screened", "", ""], ["Title/abstract excluded", "", ""],
        ["Full texts assessed", "", ""], ["Full texts excluded", "", ""], ["Studies included qualitatively", "", ""],
        ["Studies included quantitatively", "", ""],
    ])
    path = outdir / "02_srma_screening_register.xlsx"
    wb.save(path)
    return path


def extraction_workbook(outdir: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    add_note_sheet(wb, "Read me", [
        ("Gate", "Enter data only after human-confirmed inclusion and identifier verification."),
        ("Granularity", "Use one row per study in Study characteristics, and one row per outcome/timepoint/contrast elsewhere."),
        ("Source traceability", "Record table/figure/page or supplement location for every extracted numerical value."),
    ])
    add_sheet(wb, "Study characteristics", [(x, w) for x, w in [
        ("Study ID", 20), ("Citation", 20), ("PMID", 13), ("DOI", 24), ("Country/setting", 24), ("Design", 18),
        ("Population and eligibility", 42), ("Intervention/exposure", 30), ("Comparator", 28), ("Follow-up", 18), ("Funding/COI", 28), ("Verified", 12),
    ]])
    add_sheet(wb, "Outcome data", [(x, w) for x, w in [
        ("Study ID", 20), ("Outcome domain", 22), ("Outcome definition", 36), ("Timepoint", 18), ("Arm/group", 24),
        ("N randomized", 14), ("N analyzed", 14), ("Events", 12), ("Total", 12), ("Mean", 12), ("SD", 12), ("Unit/scale", 18), ("Source location", 28),
    ]])
    add_sheet(wb, "Effect estimates", [(x, w) for x, w in [
        ("Study ID", 20), ("Outcome", 24), ("Timepoint", 18), ("Effect measure", 16), ("Estimate", 14),
        ("Lower 95% CI", 15), ("Upper 95% CI", 15), ("SE", 12), ("Adjusted?", 12), ("Covariates", 32), ("Analysis set", 24), ("Source location", 28),
    ]])
    path = outdir / "03_srma_data_extraction.xlsx"
    wb.save(path)
    return path


def rob_workbook(outdir: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    add_note_sheet(wb, "Tool selection", [
        ("RCT", "Use RoB 2 (or a protocol-specified equivalent) at outcome level where appropriate."),
        ("Non-randomized intervention", "Use ROBINS-I (or a protocol-specified equivalent)."),
        ("Diagnostic/prognostic/other", "Specify the validated tool in the protocol before rating."),
        ("Rule", "Do not convert RoB judgments into certainty or pooled conclusions automatically."),
    ])
    add_sheet(wb, "RoB 2", [(x, w) for x, w in [
        ("Study ID", 20), ("Outcome/timepoint", 24), ("Randomization", 18), ("Deviations from intervention", 24),
        ("Missing outcome data", 22), ("Outcome measurement", 22), ("Selection of reported result", 28), ("Overall judgment", 20), ("Support for judgment", 48), ("Reviewer/date", 22),
    ]])
    add_sheet(wb, "ROBINS-I", [(x, w) for x, w in [
        ("Study ID", 20), ("Outcome/timepoint", 24), ("Confounding", 18), ("Selection", 18), ("Classification", 20),
        ("Deviations", 18), ("Missing data", 18), ("Outcome measurement", 22), ("Reported result", 20), ("Overall judgment", 20), ("Support", 48), ("Reviewer/date", 22),
    ]])
    path = outdir / "04_srma_risk_of_bias.xlsx"
    wb.save(path)
    return path


def meta_input_workbook(outdir: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    add_note_sheet(wb, "Read me", [
        ("Precondition", "Populate only checked numerical data from 03_srma_data_extraction.xlsx after outcome definitions and estimand are fixed."),
        ("Do not pool yet", "Choose effect measure, model, heterogeneity handling, and sensitivity analyses in the protocol/SAP before running meta-analysis."),
        ("Audit", "Keep one row per study-outcome-timepoint-contrast and record the source location."),
    ])
    add_sheet(wb, "Binary", [(x, w) for x, w in [
        ("Study ID", 20), ("Outcome", 24), ("Timepoint", 18), ("Events treatment", 16), ("Total treatment", 16),
        ("Events control", 16), ("Total control", 16), ("Effect measure", 16), ("Source location", 28),
    ]])
    add_sheet(wb, "Continuous", [(x, w) for x, w in [
        ("Study ID", 20), ("Outcome", 24), ("Timepoint", 18), ("Mean treatment", 16), ("SD treatment", 14), ("N treatment", 14),
        ("Mean control", 16), ("SD control", 14), ("N control", 14), ("Scale/unit", 18), ("Source location", 28),
    ]])
    add_sheet(wb, "Generic IV", [(x, w) for x, w in [
        ("Study ID", 20), ("Outcome", 24), ("Timepoint", 18), ("Effect measure", 16), ("log effect", 14), ("SE", 12),
        ("Adjusted?", 12), ("Covariates", 32), ("Source location", 28),
    ]])
    add_sheet(wb, "Analysis plan", [("Item", 28), ("Pre-specified decision", 70), ("Status", 18)], [
        ["Estimand / PICO", "", "TO VERIFY"], ["Effect measure by outcome", "", "TO VERIFY"], ["Model and heterogeneity", "", "TO VERIFY"],
        ["Subgroups / sensitivity analyses", "", "TO VERIFY"], ["Small-study bias assessment", "", "TO VERIFY"],
    ])
    path = outdir / "05_srma_meta_analysis_input.xlsx"
    wb.save(path)
    return path


def main() -> int:
    parser = argparse.ArgumentParser(description="Create SRMA Excel workbooks from a PubMed record set.")
    parser.add_argument("records_json", type=Path)
    parser.add_argument("outdir", type=Path)
    parser.add_argument("--query", type=Path)
    args = parser.parse_args()
    try:
        records = load_records(args.records_json)
        query = json.loads(args.query.read_text(encoding="utf-8-sig")) if args.query else {}
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        print(f"Input error: {exc}")
        return 2
    args.outdir.mkdir(parents=True, exist_ok=True)
    paths = [screening_workbook(records, args.outdir, query), extraction_workbook(args.outdir), rob_workbook(args.outdir), meta_input_workbook(args.outdir)]
    for path in paths:
        print(f"Wrote {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
